import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

# Read JSON file
with open('klaviyo-templates-metrics.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
section_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
section_text_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
filled_cell_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=14)
section_font = Font(bold=True, color="FFFFFF", size=12)
label_font = Font(bold=True, size=11)
filled_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Case Study Data for Paws Up Pet Shop
CASE_STUDY = {
    "business_name": "Paws Up Pet Shop",
    "location": "Luxembourg",
    "industry": "Pet Supplies E-commerce",
    "currency": "USD",
    "website": "https://paws-up.shop/",
    
    # Brand & Product Strategy
    "brand_personality": "Friendly",
    "total_skus": 250,
    "categories": "For Dogs: Walk, Food, Treats, Bottles/Bowls/Mats, Beds, Care, Carriers, For Car, Travel, Toys, Clothes, For Puppies\nFor Cats: Food, Treats, Beds, Care, Carriers, Toys, Clothes\nFor Owners: Accessories, Treat Pouches, Vet Passport Covers\nGrooming Services",
    "price_range": {"low": 6, "mid": 28, "high": 85},
    "bestsellers": "Snozy Bed, Cable Knitted Jumper, Reversible Puffer Jacket, Christmas Sweaters, Treats (Venison Tendon, Bulls Tail), Grooming Services",
    "new_products_frequency": "Monthly",
    
    # Personas
    "persona_1": {
        "name": "Urban Dog Parent - Sarah",
        "age": "25-34",
        "gender": "Female",
        "location": "Luxembourg City",
        "income": "Upper-middle",
        "interests": "Pet wellness, sustainable products, Instagram pet photos",
        "values": "Quality, sustainability, pet health",
        "lifestyle": "Active, urban professional, social media active",
        "frequency": "Monthly",
        "channels": "Email, Website",
        "device": "Mobile",
        "categories": "Premium treats, stylish clothes, comfortable beds",
        "price_sensitivity": "Price-neutral",
        "pain_points": "Finding quality products, keeping pet stylish and comfortable",
        "communication": "Friendly"
    },
    
    "persona_2": {
        "name": "New Puppy Owner - Michael",
        "age": "35-44",
        "gender": "Male",
        "location": "Hesperange area",
        "income": "Middle",
        "interests": "Training, pet care basics, local community",
        "values": "Education, reliability, local support",
        "lifestyle": "Family-oriented, busy schedule, values convenience",
        "frequency": "Bi-weekly",
        "channels": "Email, SMS, In-store",
        "device": "All",
        "categories": "Puppy essentials, training treats, carriers, grooming",
        "price_sensitivity": "Somewhat price-sensitive",
        "pain_points": "Knowing what products to buy, training challenges",
        "communication": "Professional"
    },
    
    # Business Intelligence
    "monthly_revenue_target": 16200,
    "email_contribution": 25,
    "repeat_purchase_rate": 35,
    "time_between_purchases": "1-2 months",
    "cart_abandonment": 68,
    "avg_abandoned_cart": 49,
    "browse_abandonment": 72,
    "current_open_rate": 22,
    "current_ctr": 3.5,
    "current_conversion": 2.8,
    
    # Competitors
    "competitor_1": {
        "name": "Zooplus",
        "frequency": "Daily",
        "email_types": "Promotional",
        "subject_style": "Direct",
        "design": "Product-focused",
        "personalization": "Moderate",
        "discount_frequency": "Frequently"
    },
    "competitor_2": {
        "name": "Fressnapf",
        "frequency": "2-3x per week",
        "email_types": "Mixed",
        "subject_style": "Benefit-focused",
        "design": "Modern",
        "personalization": "Basic (name only)",
        "discount_frequency": "Occasionally"
    }
}

def extract_dropdown_options(field_value):
    """Extract dropdown options from field value"""
    if isinstance(field_value, str):
        if field_value.startswith("Select:"):
            options = field_value.replace("Select:", "").strip().split(", ")
            return options
        elif "Yes/No select" in field_value or "Yes/No" in field_value:
            return ["Yes", "No"]
    return None

def get_filled_value(field_key, field_type, section_path=""):
    """Get pre-filled value based on field key and case study data"""
    key_lower = field_key.lower()
    
    # Checkbox fields - set to "Yes" for most, "No" for some
    if "checkbox" in field_type.lower():
        # Set strategic ones to Yes
        if any(x in key_lower for x in ["accurate", "clear", "functional", "tested", "optimized", "checked", "present", "working", "correct", "proper"]):
            return "Yes"
        return "No"
    
    # Select fields - return appropriate option
    if "select:" in field_type.lower():
        options = extract_dropdown_options(field_type)
        if options:
            # Smart selection based on context
            if "personality" in key_lower or "brand_personality" in key_lower:
                return CASE_STUDY.get("brand_personality", options[0])
            elif "frequency" in key_lower:
                if "new_product" in key_lower:
                    return CASE_STUDY.get("new_products_frequency", options[0])
                return options[1] if len(options) > 1 else options[0]
            elif "subscription" in key_lower:
                return "No"
            elif "seasonal" in key_lower:
                return "Holiday Season"
            elif "gender" in key_lower:
                return "All"
            elif "age" in key_lower:
                return "25-34"
            elif "income" in key_lower:
                return "Middle"
            elif "device" in key_lower or "preference" in key_lower:
                return "Mobile"
            elif "sensitivity" in key_lower:
                return "Price-neutral"
            elif "communication" in key_lower or "style" in key_lower:
                return "Friendly"
            elif "frequency" in key_lower and "email" in key_lower:
                return "Weekly"
            elif "type" in key_lower and "email" in key_lower:
                return "Promotional"
            elif "personalization" in key_lower:
                return "Moderate"
            elif "discount" in key_lower:
                return "Occasionally"
            elif "position" in key_lower and "market" in key_lower:
                return "Mid-market"
            elif "price_positioning" in key_lower:
                return "Similar"
            elif "technology" in key_lower:
                return "AI personalization"
            elif "language" in key_lower:
                return "English"
            elif "currency" in key_lower:
                return "USD"
            elif "loyalty" in key_lower:
                return "Moderate impact"
            elif "time_between" in key_lower:
                return "1-2 months"
            elif "time_to_abandon" in key_lower:
                return "6-24 hours"
            elif "benchmark" in key_lower:
                return "At industry average"
            return options[0]
    
    # Text inputs with specific values
    if "campaign_flow_name" in key_lower or "campaign_name" in key_lower or "flow_name" in key_lower:
        if "welcome" in section_path.lower():
            return "Welcome Series - New Pet Parents"
        elif "abandoned" in section_path.lower():
            return "Abandoned Cart Recovery - Q1 2025"
        elif "post" in section_path.lower() and "purchase" in section_path.lower():
            return "Post-Purchase Follow-up"
        elif "re-engagement" in section_path.lower():
            return "Re-Engagement Campaign - Inactive Customers"
        return "Holiday Sale Campaign - December 2024"
    
    if "segment_name" in key_lower:
        return "High-Value Repeat Customers"
    
    if "subject_line" in key_lower:
        return "🎄 20% Off Holiday Collection - Limited Time!"
    
    if "preheader" in key_lower:
        return "Treat your furry friend this holiday season"
    
    if "goal_objective" in key_lower:
        return "Increase holiday sales by 30%, drive repeat purchases, improve customer lifetime value"
    
    if "target_segment" in key_lower or "target_audience" in key_lower:
        return "Active customers (purchased in last 90 days), repeat buyers, high-value customers"
    
    if "sender_name" in key_lower:
        return "Paws Up Pet Shop"
    
    if "sender_email" in key_lower:
        return "hello@paws-up.shop"
    
    if "launch_date" in key_lower or "send_date" in key_lower:
        return (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    
    # Number inputs
    if "total_skus" in key_lower:
        return CASE_STUDY["total_skus"]
    if "low_price" in key_lower:
        return CASE_STUDY["price_range"]["low"]
    if "mid_price" in key_lower:
        return CASE_STUDY["price_range"]["mid"]
    if "high_price" in key_lower:
        return CASE_STUDY["price_range"]["high"]
    if "monthly_revenue" in key_lower:
        return CASE_STUDY["monthly_revenue_target"]
    if "quarterly_revenue" in key_lower:
        return CASE_STUDY["monthly_revenue_target"] * 3
    if "annual_revenue" in key_lower:
        return CASE_STUDY["monthly_revenue_target"] * 12
    if "emails_sent" in key_lower:
        return 5000
    if "emails_delivered" in key_lower:
        return 4850
    if "segment_size" in key_lower or "size_of_segment" in key_lower:
        return "2,500 recipients"
    if "active_subscribers" in key_lower:
        return 3500
    if "inactive" in key_lower and "subscribers" in key_lower:
        return 1200
    if "suppressed" in key_lower:
        return 150
    
    # Percentage inputs
    if "open_rate" in key_lower and "target" not in key_lower and "average" not in key_lower:
        return f"{CASE_STUDY['current_open_rate']}%"
    if "ctr" in key_lower or "click" in key_lower and "rate" in key_lower:
        if "target" in key_lower:
            return "4.5%"
        return f"{CASE_STUDY['current_ctr']}%"
    if "conversion_rate" in key_lower:
        if "target" in key_lower:
            return "3.5%"
        return f"{CASE_STUDY['current_conversion']}%"
    if "delivery_rate" in key_lower:
        return "97%"
    if "bounce_rate" in key_lower:
        return "2.1%"
    if "hard_bounce" in key_lower:
        return "0.8%"
    if "soft_bounce" in key_lower:
        return "1.3%"
    if "spam_complaint" in key_lower:
        return "0.05%"
    if "unsubscribe_rate" in key_lower:
        return "0.2%"
    if "cart_abandonment" in key_lower:
        return f"{CASE_STUDY['cart_abandonment']}%"
    if "browse_abandonment" in key_lower:
        return f"{CASE_STUDY['browse_abandonment']}%"
    if "repeat_purchase" in key_lower and "rate" in key_lower:
        return f"{CASE_STUDY['repeat_purchase_rate']}%"
    if "recovery_rate" in key_lower:
        return "12%"
    if "revenue_contribution" in key_lower:
        return f"{CASE_STUDY['email_contribution']}%"
    if "growth_rate" in key_lower:
        return "15%"
    
    # Currency inputs
    if "revenue" in key_lower and ("rpr" in key_lower or "per_recipient" in key_lower):
        return "$20.00"
    if "aov" in key_lower or "average_order" in key_lower:
        return "$56.00"
    if "total_revenue" in key_lower or "total_sales" in key_lower:
        return "$4,600"
    if "revenue_per_email" in key_lower:
        return "$0.92"
    if "abandoned_cart_value" in key_lower:
        return f"${CASE_STUDY['avg_abandoned_cart']}"
    
    # Textarea fields with realistic content
    if "textarea" in field_type.lower():
        if "mission" in key_lower or "values" in key_lower:
            return "Provide high-quality, sustainable pet products that enhance the lives of pets and their owners. We believe in ethical sourcing, community support, and exceptional customer service."
        if "messaging" in key_lower or "pillars" in key_lower:
            return "Quality products, Pet wellness, Community support, Sustainable practices, Personalized service"
        if "differentiators" in key_lower:
            return "Local Luxembourg presence with physical store, Expert grooming services, Curated product selection, Personalized customer service, Strong community focus"
        if "story" in key_lower or "narrative" in key_lower:
            return "Founded by pet lovers for pet lovers. We started as a small local shop and grew into a trusted destination for quality pet supplies, combining online convenience with in-store expertise."
        if "categories" in key_lower:
            return CASE_STUDY["categories"]
        if "bestsellers" in key_lower:
            return CASE_STUDY["bestsellers"]
        if "strategy" in key_lower:
            return "Focus on seasonal collections (holiday, summer), highlight new arrivals monthly, promote grooming services, emphasize quality and sustainability"
        if "opportunities" in key_lower:
            return "Cross-sell: Treats with toys, Beds with blankets, Carriers with travel accessories. Upsell: Premium food lines, Designer clothing, Professional grooming packages"
        if "insights" in key_lower or "analysis" in key_lower or "recommendations" in key_lower:
            return "Customers respond well to product bundles, seasonal themes drive engagement, grooming service reminders increase repeat visits, personalized product recommendations based on pet type perform 2x better"
        if "worked_well" in key_lower or "worked_best" in key_lower:
            return "Subject lines with emojis increased opens by 15%, personalized product recommendations, clear CTAs with urgency, mobile-optimized design, holiday-themed content"
        if "underperformed" in key_lower:
            return "Generic promotional emails, too frequent sends (daily), text-heavy emails without images, non-personalized content"
        if "next_steps" in key_lower or "actions" in key_lower:
            return "Implement A/B testing for subject lines, create more personalized segments, optimize send times based on timezone, develop post-purchase flows"
        if "notes" in key_lower or "observations" in key_lower:
            return "Holiday season shows 40% increase in engagement. Mobile opens are 65% of total. Repeat customers have 3x higher conversion rate."
        if "rules" in key_lower or "criteria" in key_lower:
            return "Purchased 2+ times in last 6 months, AOV > $50, Opened email in last 30 days, Located in Luxembourg"
        if "exclusion" in key_lower:
            return "Exclude: Unsubscribed, Suppressed, Bounced emails, Disengaged 90+ days"
        if "reasoning" in key_lower or "logic" in key_lower:
            return "Based on historical data, customers who purchase during holiday season have 60% higher LTV. Personalized recommendations increase conversion by 2.5x."
        if "expected" in key_lower or "hypothesis" in key_lower:
            return "Personalized subject lines will increase open rate by 10%, Product recommendations will boost CTR by 15%, Holiday-themed content will drive 30% more revenue"
        if "persona_name" in key_lower:
            return CASE_STUDY["persona_1"]["name"]
        if "interests" in key_lower:
            return CASE_STUDY["persona_1"]["interests"]
        if "values" in key_lower and "persona" in section_path.lower():
            return CASE_STUDY["persona_1"]["values"]
        if "lifestyle" in key_lower:
            return CASE_STUDY["persona_1"]["lifestyle"]
        if "pain_points" in key_lower or "motivations" in key_lower:
            return CASE_STUDY["persona_1"]["pain_points"]
        if "engagement_patterns" in key_lower:
            return "Opens emails 2-3x per week, Clicks on product links, Purchases within 48 hours of email, Shares products on social media"
        if "journey" in key_lower or "stage" in key_lower:
            if "awareness" in key_lower:
                return "Discovers through Instagram ads, Google search, word-of-mouth, local community events"
            if "consideration" in key_lower:
                return "Reads product reviews, compares prices, checks availability, visits physical store"
            if "purchase" in key_lower:
                return "Price, quality, availability, shipping options, return policy"
            if "post" in key_lower:
                return "Leaves reviews, shares on social, returns for grooming, subscribes to newsletter"
            if "retention" in key_lower:
                return "Seasonal purchases, grooming reminders, new product launches, loyalty rewards"
            if "advocacy" in key_lower:
                return "Refers friends, writes reviews, follows on social media, participates in community events"
        if "behavior" in key_lower or "patterns" in key_lower:
            return "Peak purchasing: November-December (holidays), March (spring), June (summer). Average 1.5 purchases per quarter. Prefers bundles and seasonal collections."
        if "lifecycle" in key_lower or "segments" in key_lower:
            if "new" in key_lower:
                return "First-time customers, need onboarding, welcome series, product education"
            if "active" in key_lower:
                return "Regular purchasers, engaged with emails, high open rates, seasonal buyers"
            if "repeat" in key_lower:
                return "3+ purchases, loyal customers, high AOV, refer friends"
            if "vip" in key_lower or "high_value" in key_lower:
                return "Top 20% by revenue, purchase monthly, high engagement, brand advocates"
            if "at_risk" in key_lower:
                return "60-90 days inactive, declining engagement, need re-engagement campaign"
            if "dormant" in key_lower:
                return "90+ days inactive, low email engagement, win-back campaign needed"
            if "win_back" in key_lower:
                return "Special offers, new product launches, exclusive discounts, survey to understand why they left"
        if "geographic" in key_lower or "regional" in key_lower:
            return "Top regions: Luxembourg City (45%), Hesperange area (25%), Esch-sur-Alzette (15%), Other Luxembourg (15%). Peak send times: 9-10 AM, 6-7 PM CET"
        if "competitor" in key_lower or "positioning" in key_lower:
            return "Zooplus: Large selection, frequent discounts, less personal. Fressnapf: Established brand, moderate personalization. Our advantage: Local expertise, personalized service, community focus"
        if "advantages" in key_lower or "better" in key_lower:
            return "Local physical store for try-before-buy, Expert grooming services, Curated quality selection, Personalized recommendations, Strong community presence, Sustainable product focus"
        if "trends" in key_lower:
            return "AI-powered personalization, Interactive emails, Video content, Sustainability focus, Mobile-first design, Omnichannel experience"
        if "improvements" in key_lower:
            return "Reduce email frequency to 2-3x per week, Add more visual content, Improve mobile experience, Better segmentation, A/B test subject lines"
        if "ab_testing" in key_lower or "test_ideas" in key_lower:
            return "Subject line: Emoji vs no emoji, CTA: Button text variations, Send time: Morning vs evening, Content: Product-focused vs lifestyle, Personalization: Name vs product recommendations"
        if "content" in key_lower or "copy" in key_lower or "message" in key_lower:
            return "Highlight quality and sustainability, Emphasize local community support, Showcase happy pets, Include customer testimonials, Clear value propositions"
        if "cta" in key_lower:
            return "Shop Now, View Collection, Book Grooming, Get 10% Off, See New Arrivals"
        if "dynamic" in key_lower or "personalization" in key_lower:
            return "Customer name, Pet name (if provided), Recently viewed products, Recommended products based on purchase history, Location-based offers"
        if "images" in key_lower or "media" in key_lower:
            return "High-quality product photos, Lifestyle images with pets, Seasonal themed visuals, Customer photos, Grooming service images"
        if "review" in key_lower or "approval" in key_lower:
            return "Content reviewed by marketing manager, Approved by store owner, Legal compliance checked, Brand guidelines verified"
        if "reporting" in key_lower:
            return "Weekly performance reports, Monthly revenue analysis, Segment performance tracking, A/B test results, Recommendations for optimization"
        return "See case study notes and recommendations for detailed information."
    
    # Default empty for text inputs
    return ""

def create_sheet_for_template(template_key, template_data):
    """Create a sheet for a template with filled data"""
    sheet_name = template_data['name'].replace('/', ' ').replace('\\', ' ').replace('?', '').replace('*', '').replace('[', '').replace(']', '')[:31]
    ws = wb.create_sheet(title=sheet_name)
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = f"{template_data['name']} - {CASE_STUDY['business_name']} Case Study"
    cell.font = Font(bold=True, size=18, color="FF6B35")
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 2
    
    # Process metrics
    for section_key, section_data in template_data['metrics'].items():
        section_path = f"{template_key}.{section_key}"
        # Section Header
        section_name = section_key.replace('_', ' ').title()
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = section_name
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        row += 1
        
        # Process fields in section
        process_fields(ws, section_data, row, 1, section_path)
        row = ws.max_row + 2
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

def process_fields(ws, data, start_row, col, section_path="", level=0):
    """Recursively process fields and fill with case study data"""
    current_row = start_row
    
    for key, value in data.items():
        current_path = f"{section_path}.{key}" if section_path else key
        
        if isinstance(value, dict):
            # Subsection
            subsection_name = key.replace('_', ' ').title()
            ws.merge_cells(f'{get_column_letter(col)}{current_row}:{get_column_letter(col+2)}{current_row}')
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.value = subsection_name
            cell.fill = section_text_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            current_row += 1
            
            current_row = process_fields(ws, value, current_row, col, current_path, level+1)
        elif isinstance(value, list):
            # Checklist items - convert to Yes/No dropdowns with filled values
            for item in value:
                # Label
                cell = ws[f'{get_column_letter(col)}{current_row}']
                cell.value = str(item)
                cell.font = label_font
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top')
                
                # Value/Input cell with Yes/No dropdown
                value_cell = ws[f'{get_column_letter(col+1)}{current_row}']
                value_cell.border = border
                value_cell.alignment = Alignment(horizontal='left', vertical='top')
                
                # Fill with "Yes" for most checklist items
                filled_value = "Yes" if any(x in str(item).lower() for x in ["tested", "checked", "optimized", "correct", "present", "working"]) else "No"
                value_cell.value = filled_value
                value_cell.fill = filled_cell_fill
                value_cell.font = filled_font
                
                # Add dropdown
                dv = DataValidation(type="list", formula1='"Yes,No"')
                dv.error = 'Invalid option'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Select: Yes or No'
                ws.add_data_validation(dv)
                dv.add(value_cell)
                
                # Type indicator
                type_cell = ws[f'{get_column_letter(col+2)}{current_row}']
                type_cell.value = "Yes/No Dropdown"
                type_cell.font = Font(size=9, italic=True, color="666666")
                type_cell.border = border
                type_cell.alignment = Alignment(horizontal='left', vertical='top')
                
                current_row += 1
        else:
            # Regular field
            field_label = key.replace('_', ' ').title()
            field_type = str(value)
            
            # Label
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.value = field_label
            cell.font = label_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top')
            
            # Value/Input cell
            value_cell = ws[f'{get_column_letter(col+1)}{current_row}']
            value_cell.border = border
            value_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Get filled value
            filled_value = get_filled_value(key, field_type, current_path)
            
            # Check for dropdown
            dropdown_options = extract_dropdown_options(field_type)
            if dropdown_options:
                # Add dropdown
                options_str = ",".join(dropdown_options)
                dv = DataValidation(type="list", formula1=f'"{options_str}"')
                dv.error = 'Invalid option'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = f'Select from: {", ".join(dropdown_options)}'
                ws.add_data_validation(dv)
                dv.add(value_cell)
                if filled_value:
                    value_cell.value = filled_value
                    value_cell.fill = filled_cell_fill
                    value_cell.font = filled_font
            elif "Checkbox" in field_type or "Yes/No" in field_type:
                # Convert checkbox to Yes/No dropdown
                dv = DataValidation(type="list", formula1='"Yes,No"')
                dv.error = 'Invalid option'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Select: Yes or No'
                ws.add_data_validation(dv)
                dv.add(value_cell)
                if filled_value:
                    value_cell.value = filled_value
                    value_cell.fill = filled_cell_fill
                    value_cell.font = filled_font
            else:
                # Regular text/number field
                if filled_value:
                    value_cell.value = filled_value
                    value_cell.fill = filled_cell_fill
                    value_cell.font = filled_font
            
            # Type indicator
            type_cell = ws[f'{get_column_letter(col+2)}{current_row}']
            type_cell.value = field_type.split(" -")[0] if " -" in field_type else field_type
            type_cell.font = Font(size=9, italic=True, color="666666")
            type_cell.border = border
            type_cell.alignment = Alignment(horizontal='left', vertical='top')
            
            current_row += 1
    
    return current_row

# Create sheets for each template in order
template_order = [
    "ecommerce_strategic_foundation_template",
    "email_campaign_template",
    "flow_planning_template",
    "segmentation_template",
    "success_measurement_template",
    "email_reputation_deliverability_template",
    "email_marketing_compliance_template",
    "content_copy_checklist_template",
    "campaign_post_mortem_template"
]

for template_key in template_order:
    if template_key in data['klaviyo_templates']:
        create_sheet_for_template(template_key, data['klaviyo_templates'][template_key])

# Save workbook
output_file = f'Paws_Up_Pet_Shop_Case_Study_{datetime.now().strftime("%Y%m%d")}.xlsx'
wb.save(output_file)
print(f"✅ Case Study Excel file created successfully: {output_file}")
print(f"📊 Total sheets: {len(wb.sheetnames)}")
print(f"🏢 Business: {CASE_STUDY['business_name']}")
print(f"🌍 Location: {CASE_STUDY['location']}")
print(f"📋 Sheets: {', '.join(wb.sheetnames)}")

