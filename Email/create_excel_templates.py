import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Read JSON file
with open('klaviyo-templates-metrics.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
section_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
section_text_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=14)
section_font = Font(bold=True, color="FFFFFF", size=12)
label_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def extract_dropdown_options(field_value):
    """Extract dropdown options from field value"""
    if isinstance(field_value, str):
        if field_value.startswith("Select:"):
            options = field_value.replace("Select:", "").strip().split(", ")
            return options
        elif "Yes/No select" in field_value or "Yes/No" in field_value:
            return ["Yes", "No"]
    return None

def create_sheet_for_template(template_key, template_data):
    """Create a sheet for a template"""
    # Clean sheet name (remove invalid characters)
    sheet_name = template_data['name'].replace('/', ' ').replace('\\', ' ').replace('?', '').replace('*', '').replace('[', '').replace(']', '')[:31]
    ws = wb.create_sheet(title=sheet_name)  # Excel sheet name limit
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws[f'A{row}']
    cell.value = template_data['name']
    cell.font = Font(bold=True, size=18, color="FF6B35")
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 2
    
    # Process metrics
    for section_key, section_data in template_data['metrics'].items():
        # Section Header
        section_name = section_key.replace('_', ' ').title()
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = section_name
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        row += 1
        
        # Process fields in section
        process_fields(ws, section_data, row, 1)
        row = ws.max_row + 2
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15

def process_fields(ws, data, start_row, col, level=0):
    """Recursively process fields"""
    current_row = start_row
    
    for key, value in data.items():
        if isinstance(value, dict):
            # Subsection
            subsection_name = key.replace('_', ' ').title()
            ws.merge_cells(f'{get_column_letter(col)}{current_row}:{get_column_letter(col+2)}{current_row}')
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.value = subsection_name
            cell.fill = section_text_fill
            cell.font = label_font
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            current_row += 1
            
            # Process nested fields
            current_row = process_fields(ws, value, current_row, col, level+1)
        elif isinstance(value, list):
            # Checklist items - convert to Yes/No dropdowns
            for item in value:
                # Label
                cell = ws[f'{get_column_letter(col)}{current_row}']
                cell.value = str(item)
                cell.font = label_font
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top')
                
                # Value/Input cell with Yes/No dropdown
                value_cell = ws[f'{get_column_letter(col+1)}{current_row}']
                value_cell.border = border
                value_cell.alignment = Alignment(horizontal='left', vertical='top')
                
                # Add Yes/No dropdown
                dv = DataValidation(type="list", formula1='"Yes,No"')
                dv.error = 'Invalid option'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Select: Yes or No'
                ws.add_data_validation(dv)
                dv.add(value_cell)
                
                # Type indicator
                type_cell = ws[f'{get_column_letter(col+2)}{current_row}']
                type_cell.value = "Yes/No Dropdown"
                type_cell.font = Font(size=9, italic=True, color="666666")
                type_cell.border = border
                type_cell.alignment = Alignment(horizontal='left', vertical='top')
                
                current_row += 1
        else:
            # Regular field
            field_label = key.replace('_', ' ').title()
            field_type = str(value)
            
            # Label
            cell = ws[f'{get_column_letter(col)}{current_row}']
            cell.value = field_label
            cell.font = label_font
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top')
            
            # Value/Input cell
            value_cell = ws[f'{get_column_letter(col+1)}{current_row}']
            value_cell.border = border
            value_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Check for dropdown (including Yes/No selects)
            dropdown_options = extract_dropdown_options(field_type)
            if dropdown_options:
                # Add dropdown - use comma-separated list
                options_str = ",".join(dropdown_options)
                dv = DataValidation(type="list", formula1=f'"{options_str}"')
                dv.error = 'Invalid option'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = f'Select from: {", ".join(dropdown_options)}'
                ws.add_data_validation(dv)
                dv.add(value_cell)
                # Don't set default value, leave empty
            elif "Checkbox" in field_type or "Yes/No" in field_type:
                # Convert checkbox to Yes/No dropdown
                dv = DataValidation(type="list", formula1='"Yes,No"')
                dv.error = 'Invalid option'
                dv.errorTitle = 'Invalid Entry'
                dv.prompt = 'Select: Yes or No'
                ws.add_data_validation(dv)
                dv.add(value_cell)
                # Leave empty, user will select
            elif "Textarea" in field_type:
                value_cell.value = ""
            elif "Number input" in field_type:
                value_cell.value = ""
            elif "Date input" in field_type:
                value_cell.value = ""
            else:
                value_cell.value = ""
            
            # Type indicator
            type_cell = ws[f'{get_column_letter(col+2)}{current_row}']
            type_cell.value = field_type.split(" -")[0] if " -" in field_type else field_type
            type_cell.font = Font(size=9, italic=True, color="666666")
            type_cell.border = border
            type_cell.alignment = Alignment(horizontal='left', vertical='top')
            
            current_row += 1
    
    return current_row

# Create sheets for each template in order
template_order = [
    "ecommerce_strategic_foundation_template",
    "email_campaign_template",
    "flow_planning_template",
    "segmentation_template",
    "success_measurement_template",
    "email_reputation_deliverability_template",
    "email_marketing_compliance_template",
    "content_copy_checklist_template",
    "campaign_post_mortem_template"
]

for template_key in template_order:
    if template_key in data['klaviyo_templates']:
        create_sheet_for_template(template_key, data['klaviyo_templates'][template_key])

# Save workbook
output_file = 'Klaviyo_Email_Marketing_Strategic_Templates.xlsx'
try:
    wb.save(output_file)
except PermissionError:
    # If file is open, save with timestamp
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f'Klaviyo_Email_Marketing_Strategic_Templates_{timestamp}.xlsx'
    wb.save(output_file)
    print(f"⚠️  Original file is open. Saved as: {output_file}")
print(f"✅ Excel file created successfully: {output_file}")
print(f"📊 Total sheets: {len(wb.sheetnames)}")
print(f"📋 Sheets: {', '.join(wb.sheetnames)}")

