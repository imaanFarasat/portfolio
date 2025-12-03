import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Nov-Dec 2025 Calendar"

# Define colors
header_fill = PatternFill(start_color="FF8B4513", end_color="FF8B4513", fill_type="solid")  # Brown/Orange
header_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")  # Light orange
border_style = Side(style="thin", color="000000")
border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

# Headers
headers = ["Date", "Day", "Channel", "Campaign Type", "Audience / Segment", "Subject Line / Hook", "CTA", "Notes"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Calendar data for November-December 2025
# Note: Black Friday 2025 is Nov 28, Cyber Monday 2025 is Dec 1
calendar_data = [
    ["Nov 1, 2025", "Saturday", "Email", "Welcome / Onboarding", "New subscribers", "Welcome! Here's 10% off your first order", "Explore Collection", "Start-of-month welcome flow"],
    ["Nov 5, 2025", "Wednesday", "SMS", "Flash Sale", "Engaged customers", "48hr Flash Sale – Don't Miss Out!", "Grab Your Deal", "Short, urgent SMS"],
    ["Nov 10, 2025", "Monday", "Email", "Product Launch", "All subscribers", "Meet our New Winter Collection ❄️", "See What's New", "Seasonal campaign"],
    ["Nov 15, 2025", "Saturday", "Email", "Educational / Value", "Inactive / cold", "How to style your winter accessories", "Get Inspired", "Content-focused to re-engage"],
    ["Nov 20, 2025", "Thursday", "Email", "Black Friday Teaser", "VIP / High-value", "Get Ready: Our Black Friday Deals Are Coming", "Preview the Deals", "Build anticipation"],
    ["Nov 23, 2025", "Sunday", "SMS", "Black Friday Alert", "All subscribers", "Black Friday starts NOW! 🖤", "Unlock Savings", "High urgency"],
    ["Nov 28, 2025", "Friday", "Email", "Black Friday Sale", "All subscribers", "Black Friday Deals – Up to 50% Off", "See Offers", "Major discount day"],
    ["Nov 29, 2025", "Saturday", "Email", "Black Friday Reminder", "Cart abandoners", "Your favorites are almost gone!", "Complete Your Order", "Recovery / scarcity tactic"],
    ["Dec 1, 2025", "Monday", "Email", "Cyber Monday Sale", "All subscribers", "Cyber Monday – Last Chance to Save!", "Discover Deals", "High urgency"],
    ["Dec 1, 2025", "Monday", "Email", "Cyber Monday Teaser", "VIP", "Cyber Monday is almost here!", "Sneak a Peek", "Exclusive VIP early access"],
    ["Dec 1, 2025", "Monday", "Email", "Holiday Gift Guide", "All subscribers", "Gift Ideas for Everyone on Your List 🎁", "Browse Gift Ideas", "Holiday-focused content"],
    ["Dec 5, 2025", "Friday", "SMS", "Holiday Promo", "Engaged", "Our Holiday Sale is Live! 🎄", "Check It Out", "Short, urgent reminder"],
    ["Dec 10, 2025", "Wednesday", "Email", "New Product Highlight", "All subscribers", "Just In: Cozy Winter Essentials", "Explore New Arrivals", "Focus on new arrivals"],
    ["Dec 15, 2025", "Monday", "Email", "Last-Minute Gift Ideas", "All subscribers", "Last-Minute Gifts That Will Impress 🎁", "Find the Perfect Gift", "Time-sensitive, drives urgency"],
    ["Dec 20, 2025", "Saturday", "Email", "Shipping Deadline Reminder", "All subscribers", "Order by Dec 22 for Christmas Delivery", "See Shipping Options", "Logistics-focused"],
    ["Dec 23, 2025", "Tuesday", "SMS", "Final Holiday Reminder", "All subscribers", "Last chance for Christmas delivery!", "Check Delivery", "Urgent SMS"],
    ["Dec 26, 2025", "Friday", "Email", "Boxing Day Sale", "All subscribers", "Boxing Day Deals – Up to 60% Off", "Explore Deals", "Post-holiday promotion"],
    ["Dec 31, 2025", "Wednesday", "Email", "New Year / Wrap-Up", "All subscribers", "Cheers to 2026! 🎉 Special Offer Inside", "Celebrate With Us", "End-of-year engagement"],
]

# Add data rows
for row_num, row_data in enumerate(calendar_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        
        # Alternate row colors for better readability
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FFF9F0", end_color="FFF9F0", fill_type="solid")

# Adjust column widths
column_widths = {
    "A": 15,  # Date
    "B": 12,  # Day
    "C": 10,  # Channel
    "D": 22,  # Campaign Type
    "E": 20,  # Audience / Segment
    "F": 45,  # Subject Line / Hook
    "G": 18,  # CTA
    "H": 35,  # Notes
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Set row height for header
ws.row_dimensions[1].height = 30

# Set row heights for data rows
for row_num in range(2, len(calendar_data) + 2):
    ws.row_dimensions[row_num].height = 40

# Freeze the header row
ws.freeze_panes = "A2"

# Add title
ws.insert_rows(1)
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "📅 November – December 2025 Email/SMS Calendar (Alternative CTAs)"
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill(start_color="FF8B4513", end_color="FF8B4513", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# Adjust header row number
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Update freeze panes
ws.freeze_panes = "A3"

# Save the file
filename = "Email_Marketing_Calendar_Nov_Dec_2025.xlsx"
wb.save(filename)
print(f"✅ Calendar created successfully: {filename}")

